import openpyxl
import re

# Cargar workbook
wb = openpyxl.load_workbook('088.xlsx')
ws = wb.active

print("=== EXTRAYENDO CAMPOS DEL FORMULARIO REVISIÓN ENERGÉTICA 088 ===\n")

# Extraer todos los valores de la primera columna que tengan contenido significativo
fields = {}
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row, 1)
    value = cell.value
    
    if value and isinstance(value, str) and len(value.strip()) > 0:
        # Solo mostrar líneas que parecen ser etiquetas
        if any(keyword in value.lower() for keyword in ['fecha', 'regional', 'centro', 'sede', 'dirección', 'ciudad', 'área', 'temperatura', 'velocidad', 'radiación', 'año', 'número', 'propiedad', 'trabajadores', 'aprendices', 'visitantes', 'cuenta', 'generación', 'ubicación', 'consumo', 'energía', 'gas', 'acpm', 'gasolina']):
            print(f"Fila {row}: {value[:100]}")

print("\n=== BUSCANDO DATOS EN COLUMNAS ===\n")
data_rows = {}
for row in range(15, min(60, ws.max_row)):
    row_label = ws.cell(row, 1).value
    if row_label and isinstance(row_label, str):
        values = []
        for col in range(2, 10):
            val = ws.cell(row, col).value
            if val is not None:
                values.append(f"Col{col}: {val}")
        if values:
            print(f"Fila {row} ({row_label[:30]}): {', '.join(values)}")
