import openpyxl
from openpyxl.utils import get_column_letter

# Cargar workbook
wb = openpyxl.load_workbook('088.xlsx')
ws = wb.active

print(f"Dimensiones de la hoja: {ws.max_row} filas x {ws.max_column} columnas")
print(f"\n=== PRIMERAS 30 FILAS Y PRIMERAS 20 COLUMNAS ===\n")

# Mostrar datos de forma legible
for row in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(21, ws.max_column + 1)):
        cell = ws.cell(row, col)
        value = cell.value if cell.value is not None else ""
        row_data.append(str(value)[:15])  # Limitar a 15 caracteres por celda
    print(" | ".join(f"{v:15}" for v in row_data))

print(f"\n=== INFORMACIÓN DE GRÁFICOS ===\n")
if hasattr(ws, '_charts'):
    print(f"Gráficos encontrados: {len(ws._charts)}")
    for i, chart in enumerate(ws._charts):
        print(f"Gráfico {i+1}:")
        print(f"  - Tipo: {type(chart).__name__}")
        print(f"  - Título: {chart.title}")
        try:
            print(f"  - Posición: {chart.anchor}")
        except:
            pass
else:
    print("No hay gráficos en esta hoja")
